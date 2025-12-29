"""Servicio de exportación a múltiples formatos"""
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Servicio para exportar datos a Excel, CSV, JSON"""
    
    @staticmethod
    def exportar_estudiantes_excel(estudiantes, seguimientos_dict):
        """Exporta estudiantes a Excel con información de riesgo"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ["Código", "Nombres", "Apellidos", "Email", "Teléfono", 
                   "Categoría Riesgo", "Puntaje Riesgo", "Asistencia", "Promedio"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos
        for row, estudiante in enumerate(estudiantes, 2):
            seguimiento = seguimientos_dict.get(estudiante.id)
            
            datos = [
                estudiante.codigo_estudiante,
                estudiante.nombres,
                estudiante.apellidos,
                estudiante.email,
                estudiante.telefono or "",
                seguimiento.categoria_riesgo if seguimiento else "SIN_RIESGO",
                float(seguimiento.puntaje_riesgo) if seguimiento else 0.0,
                seguimiento.factores_riesgo.get('asistencia', 0) if seguimiento and seguimiento.factores_riesgo else 0,
                seguimiento.factores_riesgo.get('promedio_calificaciones', 0) if seguimiento and seguimiento.factores_riesgo else 0,
            ]
            
            for col, valor in enumerate(datos, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = valor
                cell.border = border
                if col >= 6:  # Columnas numéricas
                    cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def exportar_estudiantes_csv(estudiantes, seguimientos_dict):
        """Exporta estudiantes a CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        headers = ["Código", "Nombres", "Apellidos", "Email", "Teléfono", 
                   "Categoría Riesgo", "Puntaje Riesgo", "Asistencia", "Promedio"]
        writer.writerow(headers)
        
        # Datos
        for estudiante in estudiantes:
            seguimiento = seguimientos_dict.get(estudiante.id)
            
            datos = [
                estudiante.codigo_estudiante,
                estudiante.nombres,
                estudiante.apellidos,
                estudiante.email,
                estudiante.telefono or "",
                seguimiento.categoria_riesgo if seguimiento else "SIN_RIESGO",
                float(seguimiento.puntaje_riesgo) if seguimiento else 0.0,
                seguimiento.factores_riesgo.get('asistencia', 0) if seguimiento and seguimiento.factores_riesgo else 0,
                seguimiento.factores_riesgo.get('promedio_calificaciones', 0) if seguimiento and seguimiento.factores_riesgo else 0,
            ]
            writer.writerow(datos)
        
        output.seek(0)
        return output
    
    @staticmethod
    def exportar_reportes_excel(reportes):
        """Exporta reportes a Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reportes"
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ["Tipo", "Título", "Fecha Generación", "Usuario"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row, reporte in enumerate(reportes, 2):
            datos = [
                reporte.tipo_reporte,
                reporte.titulo,
                reporte.fecha_generacion.strftime('%d/%m/%Y %H:%M'),
                reporte.usuario.username if reporte.usuario else "",
            ]
            
            for col, valor in enumerate(datos, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = valor
                cell.border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
